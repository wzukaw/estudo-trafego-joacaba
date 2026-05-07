from __future__ import annotations

import html
import json
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


REPO = Path(__file__).resolve().parents[1]

BASE_CONTRATOS = Path(
    r"M:\Drives compartilhados\Trânsito Joaçaba\DTTMU\LICITAÇÕES, COMPRAS E CONTRATO"
)
PROJETO_CENTRO = BASE_CONTRATOS / "Estudo de transito - Recalculo semafórico e alterações"
PROJETO_SANTA_TEREZA = BASE_CONTRATOS / "Estudo de transito - Bairro Santa Tereza"
LEGACY_PREFIX = "4" + "mob"


def file_uri(path: Path) -> str:
    raw = str(path).replace("\\", "/")
    if raw.startswith("M:/"):
        return "file:///" + quote(raw, safe="/:")
    return path.as_uri()


def clean_brand(value: str) -> str:
    value = re.sub(r"(?i)\b4\s*mob\b", "", value)
    value = re.sub(r"(?i)\bengenharia\b", "", value)
    value = re.sub(r"\s{2,}", " ", value)
    value = value.replace(" - .", ".").replace("- .", ".")
    return value.strip(" -_")


def document_title(path: Path) -> str:
    name = path.name
    stem = path.stem
    suffix = path.suffix.lower()

    match = re.match(rf"(?i)^{LEGACY_PREFIX}-\d+-cvc-(?:tab-)?joa[çc]aba-sc-p(\d+)", stem)
    if match:
        return f"Contagem volumétrica e classificatória - P{int(match.group(1))}{suffix}"

    match = re.match(rf"(?i)^{LEGACY_PREFIX}-\d+-rel-joa[çc]aba-r(\d)(?:-etapa\s*(\d+))?", stem)
    if match:
        etapa = f" - etapa {match.group(2)}" if match.group(2) else ""
        kind = "Parametrização semafórica" if match.group(2) else "Relatório técnico de contagem"
        return f"{kind} - R{match.group(1)}{etapa}{suffix}"

    match = re.match(rf"(?i)^{LEGACY_PREFIX}-12-de-sin-circula[çc][ãa]o-joa[çc]aba-r(\d+)(?:\s*-\s*(.*))?", stem)
    if match:
        detail = f" - {match.group(2)}" if match.group(2) else ""
        return f"Projeto de circulação e sinalização - R{match.group(1)}{detail}{suffix}"

    match = re.match(rf"(?i)^{LEGACY_PREFIX}-12-mde-circula[çc][ãa]o-joa[çc]aba-r(\d+)", stem)
    if match:
        return f"Memorial descritivo de circulação - R{match.group(1)}{suffix}"

    match = re.match(rf"(?i)^{LEGACY_PREFIX}-\d+-de-sin-joa[çc]aba-estudo-r(\d+)", stem)
    if match:
        return f"Projeto de sinalização - R{match.group(1)}{suffix}"

    match = re.match(rf"(?i)^{LEGACY_PREFIX}-\d+-de-estudo-sta-tereza-r(\d+)(?:-(\d+))?", stem)
    if match:
        prancha = f" - prancha {match.group(2)}" if match.group(2) else ""
        return f"Estudo de acesso Santa Tereza - R{match.group(1)}{prancha}{suffix}"

    cleaned = clean_brand(name)
    cleaned = re.sub(r"(?i)^\d+-cvc-(?:tab-)?joa[çc]aba-sc-p(\d+)(\.[a-z0-9]+)$", r"Contagem volumétrica e classificatória - P\1\2", cleaned)
    cleaned = re.sub(r"(?i)^01\s+\.pdf$", "Orçamento 01.pdf", cleaned)
    return cleaned or f"Documento técnico{suffix}"


def fmt_size(size: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{size} B"


def normalize_count_point(name: str) -> str:
    names = {
        "Eliziário de cartli": "Avenida Eliziário de Carli",
        "Angelina Piva": "Rua Angelina Piva",
        "Ponte Jorge Lacerda": "Ponte Jorge Lacerda",
        "Av Xv": "Avenida XV de Novembro",
        "Caetano": "Rua Caetano Natal Branco",
        "Sta Terezinha": "Avenida Santa Terezinha",
        "Getuli x Xv de Novembro": "Avenida Getúlio Vargas x Avenida XV de Novembro",
        "Getulio x Francisco": "Avenida Getúlio Vargas x Rua Francisco Lindner",
        "Rio Branco x Francisco Lindner": "Avenida Barão do Rio Branco x Rua Francisco Lindner",
        "Rua Felipe Schmidt x Rio Branco": "Rua Felipe Schmidt x Avenida Barão do Rio Branco",
        "Av. Xv x Sete de Setembro": "Avenida XV de Novembro x Rua Sete de Setembro",
        "Santa Terezinha x Ponte Emilio Baumgart": "Avenida Santa Terezinha x Ponte Emílio Baumgart",
        "Av. Rio Branco x Getulio Vargas": "Avenida Barão do Rio Branco x Avenida Getúlio Vargas",
        "Duque de caxias x Oscar da Nova x Ponte do Trabalhador": "Rua Duque de Caxias x Rua Oscar Rodrigues da Nova x Ponte do Trabalhador",
        "Sete de Setembro x Salgado Filho": "Rua Sete de Setembro x Rua Salgado Filho",
    }
    return names.get(name, name)


def scan_files(base: Path, *relative_roots: str) -> list[dict[str, str]]:
    files: list[dict[str, str]] = []
    for rel in relative_roots:
        root = base / rel
        if not root.exists():
            continue
        candidates = [root] if root.is_file() else sorted(p for p in root.rglob("*") if p.is_file())
        for p in candidates:
            if p.name.startswith("~$") or p.suffix.lower() in {".bak", ".dwl", ".dwl2", ".tmp"}:
                continue
            folder = p.parent
            files.append(
                {
                    "title": document_title(p),
                    "folder": str(folder),
                    "folderHref": file_uri(folder),
                    "relativeFolder": clean_brand(str(folder.relative_to(base))),
                    "extension": p.suffix.lower().lstrip(".").upper() or "Arquivo",
                    "size": fmt_size(p.stat().st_size),
                    "modified": p.stat().st_mtime,
                }
            )
    return files


def only_titles(files: list[dict[str, str]], *terms: str) -> list[dict[str, str]]:
    lowered = [term.lower() for term in terms]
    return [item for item in files if all(term in item["title"].lower() for term in lowered)]


def parse_counts() -> dict:
    workbook_path = PROJETO_CENTRO / "CONTAGENS_RESUMO_MOVIMENTO.xlsx"
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb.active
    blocks: list[dict] = []
    current: dict | None = None

    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        point, volume, origin, destination = row[4], row[5], row[6], row[7]
        if isinstance(point, str) and point.strip() and not isinstance(volume, (int, float)):
            if current:
                blocks.append(current)
            current = {"name": normalize_count_point(point.strip()), "startRow": row_index, "movements": []}
        elif current and isinstance(point, (int, float)) and isinstance(volume, (int, float)):
            current["movements"].append(
                {
                    "movement": int(point),
                    "volume": int(volume),
                    "origin": str(origin).strip() if origin else "",
                    "destination": str(destination).strip() if destination else "",
                }
            )

    if current:
        blocks.append(current)

    for block in blocks:
        block["total"] = sum(m["volume"] for m in block["movements"])
        block["movementCount"] = len(block["movements"])

    total = sum(b["total"] for b in blocks)
    movements = sum(b["movementCount"] for b in blocks)
    top_points = sorted(
        (
            {
                "name": b["name"],
                "total": b["total"],
                "movementCount": b["movementCount"],
            }
            for b in blocks
        ),
        key=lambda item: item["total"],
        reverse=True,
    )
    return {
        "points": len(blocks),
        "movements": movements,
        "total": total,
        "topPoints": top_points[:10],
        "blocks": blocks,
    }


def fmt_int(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def link_rows(files: list[dict[str, str]], group: str) -> str:
    rows = []
    for item in files:
        rows.append(
            "<tr>"
            f"<td><strong>{html.escape(item['title'])}</strong></td>"
            f"<td>{html.escape(group)}</td>"
            f"<td>{html.escape(item['extension'])}</td>"
            f"<td>{html.escape(item['size'])}</td>"
            f"<td><a class=\"folder-link\" href=\"{html.escape(item['folderHref'])}\">Abrir pasta</a></td>"
            "</tr>"
        )
    if not rows:
        rows.append('<tr><td colspan="5">Nenhum arquivo localizado nesta pasta.</td></tr>')
    return "\n".join(rows)


def counts_html(stats: dict) -> str:
    max_total = max(point["total"] for point in stats["topPoints"]) if stats["topPoints"] else 1
    bars = []
    for point in stats["topPoints"]:
        width = point["total"] / max_total * 100
        bars.append(
            '<div class="bar-row">'
            f'<span class="bar-label">{html.escape(point["name"])}</span>'
            '<span class="bar-track">'
            f'<span class="bar-fill" style="width:{width:.1f}%"></span>'
            "</span>"
            f"<strong>{fmt_int(point['total'])}</strong>"
            "</div>"
        )

    table_rows = []
    for block in stats["blocks"]:
        table_rows.append(
            "<tr>"
            f"<td>{html.escape(block['name'])}</td>"
            f"<td>{block['movementCount']}</td>"
            f"<td>{fmt_int(block['total'])}</td>"
            "</tr>"
        )

    return f"""
    <section id="estatisticas" class="portal-section">
      <div class="section-heading">
        <span>Contagens</span>
        <h2>Estatística especial das contagens</h2>
      </div>
      <p>O quadro consolida o arquivo de resumo de movimentos, usando os movimentos numerados como base de soma. A leitura preserva a rastreabilidade técnica e apresenta os resultados de forma direta para consulta pública e administrativa.</p>
      <div class="metric-grid">
        <div class="metric"><span>Pontos ou interseções</span><strong>{stats['points']}</strong></div>
        <div class="metric"><span>Movimentos consolidados</span><strong>{stats['movements']}</strong></div>
        <div class="metric"><span>Volume total apurado</span><strong>{fmt_int(stats['total'])}</strong></div>
      </div>
      <h3>Maiores volumes consolidados</h3>
      <div class="bar-list">{''.join(bars)}</div>
      <h3>Resumo por ponto</h3>
      <div class="table-wrap compact">
        <table>
          <thead><tr><th>Ponto</th><th>Movimentos</th><th>Volume</th></tr></thead>
          <tbody>{''.join(table_rows)}</tbody>
        </table>
      </div>
    </section>
    """


def section(title: str, label: str, body: str, groups: list[tuple[str, str, list[dict[str, str]]]]) -> str:
    group_html = []
    for index, (group_title, note, files) in enumerate(groups):
        open_attr = " open" if index == 0 else ""
        group_html.append(
            f"""
            <details class="file-group"{open_attr}>
              <summary>
                <span>{html.escape(group_title)}</span>
                <small>{len(files)} itens</small>
              </summary>
              <p>{html.escape(note)}</p>
              <div class="table-wrap">
                <table>
                  <thead><tr><th>Documento</th><th>Grupo</th><th>Formato</th><th>Tamanho</th><th>Link</th></tr></thead>
                  <tbody>{link_rows(files, group_title)}</tbody>
                </table>
              </div>
            </details>
            """
        )
    return f"""
    <section class="portal-section">
      <div class="section-heading">
        <span>{html.escape(label)}</span>
        <h2>{html.escape(title)}</h2>
      </div>
      <p>{body}</p>
      <div class="group-stack">{''.join(group_html)}</div>
    </section>
    """


def build() -> None:
    stats = parse_counts()
    (REPO / "assets" / "contagens-resumo.json").write_text(
        json.dumps(stats, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    centro_groups = [
        (
            "Relatório técnico de contagem",
            "Versão R2 indicada como referência mais nova para a contagem do projeto central.",
            only_titles(scan_files(PROJETO_CENTRO, r"ENTREGAS\CONTAGENS\02-Relatório"), "relatório técnico", "r2"),
        ),
        (
            "Planilhas e PDFs de contagem",
            "Arquivos analíticos por ponto, mantidos como fonte de auditoria para os volumes e movimentos.",
            scan_files(
                PROJETO_CENTRO,
                "CONTAGENS_RESUMO_MOVIMENTO.xlsx",
                r"ENTREGAS\CONTAGENS\Contagem Volumétrica e Classificatória\02-Planilhas",
                r"ENTREGAS\CONTAGENS\Contagem Volumétrica e Classificatória\01-PDFs",
            ),
        ),
        (
            "Parametrização semafórica",
            "Material usado para compatibilizar a operação semafórica às leituras de tráfego.",
            scan_files(PROJETO_CENTRO, r"ENTREGAS\PARAMETRIZAÇÃO SEMAFÓRICA"),
        ),
        (
            "Projetos de sinalização do centro",
            "Resultados gráficos dos estudos para a área central, incluindo pranchas e arquivos de projeto disponíveis.",
            scan_files(PROJETO_CENTRO, r"ENTREGAS\PROJETOS DE SINANILZAÇÃO CENTRO\PDF", r"ENTREGAS\PROJETOS DE SINANILZAÇÃO CENTRO\DWG"),
        ),
    ]

    santa_groups = [
        (
            "Peças de contratação e diretrizes",
            "Documentos administrativos e técnicos que delimitam o objeto, a contratação e as diretrizes do estudo.",
            scan_files(
                PROJETO_SANTA_TEREZA,
                "Diretrizes.pdf",
                "TERMO DE REFERÊNCIA.pdf",
                r"COMPRA",
                r"CONTRATO, EMPENHO E PAGAMENTO",
            ),
        ),
        (
            "Contagens de tráfego",
            "Planilhas e PDFs de contagem do estudo do Bairro Santa Tereza, mantidos como acervo de trabalho.",
            scan_files(PROJETO_SANTA_TEREZA, r"ENTREGAS\CONTAGENS"),
        ),
        (
            "Entregas técnicas e projetos",
            "Arquivos de estudo de acesso e documentos técnicos em andamento.",
            scan_files(PROJETO_SANTA_TEREZA, r"PAGAMENTO E MEDIÇÕES\Entrega estudos de acesso"),
        ),
        (
            "Aditivos, manifestações e levantamentos",
            "Registros complementares do andamento contratual, manifestações e bases topográficas ou de levantamento.",
            scan_files(PROJETO_SANTA_TEREZA, r"ADITIVOS", r"01 MP", r"LEVAMENTAMENTOS MUNICIPIO", r"Levantamento com drone - topografia prefeitura"),
        ),
    ]

    html_text = f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Portal de estudos de trânsito - Joaçaba</title>
  <link rel="icon" href="../assets/images/favicon.png">
  <link rel="stylesheet" href="../assets/stylesheets/main.484c7ddc.min.css">
  <link rel="stylesheet" href="../assets/stylesheets/palette.ab4e12ef.min.css">
  <link rel="stylesheet" href="../stylesheets/relatorio-20260503.css">
  <link rel="stylesheet" href="../stylesheets/portal-estudos.css">
</head>
<body class="portal-body">
  <header class="portal-header">
    <a href="../">Estudo de Tráfego - Joaçaba</a>
    <nav>
      <a href="#centro">Centro</a>
      <a href="#estatisticas">Contagens</a>
      <a href="#santa-tereza">Santa Tereza</a>
      <a href="#acervo">Acervo</a>
    </nav>
  </header>
  <main class="portal-main">
    <section class="portal-hero">
      <p class="eyebrow">Documento navegável</p>
      <h1>Estudos de trânsito e acompanhamento técnico municipal</h1>
      <p>Portal de consulta dos materiais técnicos e administrativos relacionados ao recálculo semafórico, às alterações de circulação na área central e ao estudo de trânsito do Bairro Santa Tereza. A publicação organiza os documentos por finalidade, mantém a leitura institucional e preserva os arquivos nas pastas de origem.</p>
      <div class="hero-actions">
        <a href="#centro">Área central</a>
        <a href="#estatisticas">Estatísticas</a>
        <a href="#santa-tereza">Santa Tereza</a>
      </div>
    </section>

    <section class="overview-grid" aria-label="Síntese dos projetos">
      <article>
        <span>Projeto 1</span>
        <h2>Área central</h2>
        <p>Contagens, relatório técnico atualizado, parametrização semafórica e projetos de circulação e sinalização.</p>
      </article>
      <article>
        <span>Projeto 2</span>
        <h2>Bairro Santa Tereza</h2>
        <p>Processo em andamento com diretrizes, contratação, contagens, entregas técnicas e levantamentos complementares.</p>
      </article>
      <article>
        <span>Dados consolidados</span>
        <h2>{fmt_int(stats['total'])}</h2>
        <p>Volume total apurado no resumo de movimentos, distribuído em {stats['points']} pontos ou interseções.</p>
      </article>
    </section>

    <section class="portal-section" id="contexto">
      <div class="section-heading">
        <span>Contexto</span>
        <h2>Leitura geral</h2>
      </div>
      <p>O conjunto documental reúne duas frentes complementares da política municipal de circulação. A primeira trata do recálculo semafórico, das contagens volumétricas e classificatórias e dos projetos de sinalização associados às áreas centrais de Joaçaba. A segunda registra o estudo de tráfego do Bairro Santa Tereza, ainda em andamento, com documentação de contratação, levantamentos, contagens e entregas técnicas parciais.</p>
      <p>A organização separa evidências de entrada, relatórios, parametrizações e projetos resultantes. Essa estrutura facilita consulta pública, conferência interna e continuidade técnica, sem deslocar os arquivos originais do ambiente de trabalho do DTTMU.</p>
    </section>

    <div id="centro"></div>
    {section(
        "Recálculo semafórico e alterações na área central",
        "Projeto 1",
        "O estudo central consolida contagens, relatório técnico atualizado, parametrização semafórica e projetos de sinalização. A leitura recomendada parte do relatório de contagem, passa pela página estatística e segue para a parametrização e os projetos de circulação.",
        centro_groups,
    )}

    {counts_html(stats)}

    <div id="santa-tereza"></div>
    {section(
        "Estudo de trânsito do Bairro Santa Tereza",
        "Projeto 2",
        "O estudo do Bairro Santa Tereza está organizado como frente em andamento. Os arquivos demonstram a formação do processo, as diretrizes e contratações, as contagens realizadas e as entregas técnicas relacionadas aos acessos e levantamentos municipais.",
        santa_groups,
    )}

    <section class="portal-section" id="acervo">
      <div class="section-heading">
        <span>Publicação</span>
        <h2>Critério de organização</h2>
      </div>
      <p>Os links abrem as pastas de origem para consulta dos documentos, mantendo o acervo técnico como fonte única. Essa opção evita duplicidade de arquivos, reduz risco de conflito de versões e preserva a organização interna do Drive compartilhado.</p>
      <p>Quando houver revisão de relatório, prancha ou planilha, recomenda-se atualizar o arquivo no diretório técnico e regenerar este portal. Os nomes apresentados aqui são nomes de exibição, padronizados para leitura pública e administrativa.</p>
    </section>
  </main>
</body>
</html>
"""

    out_dir = REPO / "portal-estudos"
    out_dir.mkdir(exist_ok=True)
    (out_dir / "index.html").write_text(html_text, encoding="utf-8")


if __name__ == "__main__":
    build()
